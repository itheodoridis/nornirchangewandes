from netmiko import ConnectHandler, ssh_exception
from paramiko.ssh_exception import SSHException
import os
import sys
import subprocess
import openpyxl
import ipdb

from nornir import InitNornir
from nornir.plugins.tasks import networking
from nornir.plugins.functions.text import print_result
from nornir.plugins.tasks.networking import netmiko_send_config, netmiko_send_command

#this function defines the combination of tasks to be run towards the hosts in the inventory
def check_wan_int_description(task):
    #we run the show interfaces description command and then parse with ntc-templates to get structured output.
    task_info = task.run(netmiko_send_command,command_string="show interfaces description",use_textfsm=True)
    #the second interface is the gigabit0/0/1 interface, we get its description
    actual_wan_int_description = task_info[0].result[1]['descrip']
    #the expected description comes from the excel file we read in the main function
    expected_wan_int_description = task.host.get("expected_wan_int_description")
    #ipdb.set_trace() #this is a breakpoint for the debugger, now inactive
    if actual_wan_int_description != expected_wan_int_description:
        #we formulate the commands to be run
        cfg = f"interface gi0/0/1\ndescription {expected_wan_int_description}"
        #we first send the config commands and then a write command in enable mode
        task.run(netmiko_send_config,config_commands=cfg)
        task.run(netmiko_send_command,command_string="write")
        #we keep track of changes and hosts where we do the changes
        global changescount
        changescount = changescount + 1
        global changedhostlist
        changedhostlist.append(task.host.name)
    return

def main():
    #we read the hosts list from the excel file
    wbsites = openpyxl.load_workbook("mplsmaster.xlsx")
    ws = wbsites.active
    max_row = ws.max_row

    routers=[]
    for row in range (2, max_row+1):
        router=dict()
        router['site']=str(ws.cell(row= row, column = 1).value).strip()
        router['ipaddress']=str(ws.cell(row= row, column = 2).value).strip()
        router['wandes']="MPLS-OTE ("+str(ws.cell(row= row, column = 3).value).strip()+") ("+str(ws.cell(row= row, column = 4).value).strip()+")"
        routers.append(router)

    wbsites.close()

    #Nornir is initialiazed and the inventory is formed
    nr = InitNornir(config_file="config.yaml")

    #ipdb.set_trace()

    #we append a new attribute to every host in the inventory, which comes from the excel file
    for router in routers:
        node=router['site']
        nr.inventory.hosts[node]["expected_wan_int_description"]=router["wandes"]

    global changescount
    changescount = 0

    global changedhostlist
    changedhostlist = []

    #the custom task is run by calling the function
    result = nr.run(task=check_wan_int_description, num_workers=50)

    nr.close_connections()

    #we print the desired values, now part of the inventory
    for host in nr.inventory.hosts.values():
        print(host.get("expected_wan_int_description"))

    #we print how many changes were made and where
    print("changes made:",changescount)
    for host in changedhostlist:
        print(host)


if __name__ == "__main__":
    main()
